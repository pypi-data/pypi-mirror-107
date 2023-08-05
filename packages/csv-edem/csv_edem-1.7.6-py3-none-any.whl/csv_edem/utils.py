import logging
import os
from typing import List, Optional

import numpy as np
import openpyxl

logger = logging.getLogger(__name__)


def isValidCSV(file: str) -> bool:
    ext = os.path.splitext(file)[-1].lower()
    return os.path.isfile(file) and (ext == ".csv")


def saveDataToExcel(
    filename: str,
    data: np.ndarray,
    headers: List[str],
    sheetname: Optional[str] = None,
    column_widths: List[int] = None,
    align: List[str] = None,
) -> None:
    logger.debug("Save data to excel called - from utils")

    wb = openpyxl.Workbook()
    ws1 = wb.active

    if sheetname is not None:
        ws1.title = sheetname

    for i, header in enumerate(headers):
        ws1.cell(row=1, column=i + 1, value=header).font = openpyxl.styles.Font(
            bold=True
        )

        ws1.cell(row=1, column=i + 1).alignment = openpyxl.styles.Alignment(
            horizontal="right"
        )

        if column_widths is not None:
            cw = column_widths[i]
        else:
            cw = len(header) + 1

        if align is not None:
            ws1.cell(row=1, column=i + 1).alignment = openpyxl.styles.Alignment(
                horizontal=align[i]
            )

        ws1.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = cw

    for i in range(len(data[:, 0])):
        for j in range(len(data[0, :])):
            try:
                ws1.cell(row=i + 2, column=j + 1).value = float(data[i, j])
            except:
                ws1.cell(row=i + 2, column=j + 1, value=data[i, j])

            if align is not None:
                ws1.cell(row=i + 2, column=j + 1).alignment = openpyxl.styles.Alignment(
                    horizontal=align[j]
                )

    wb.save(filename=filename)

    logger.info('Exported data to excel file: "{}"'.format(filename))
