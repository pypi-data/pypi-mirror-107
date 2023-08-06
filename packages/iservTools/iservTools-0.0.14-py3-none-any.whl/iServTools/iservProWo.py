# coding: utf-8
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys

import time, datetime
import os, platform
import sqlite3, csv
import re
import iservToolBox
from iservData import iservObject

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color

import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

class ProjektwocheObject:
    """Management einer Projektwoche
    
       :param iserv: (iservObject) iserv-Objekt aus dem Modul iservData
    """
    def __init__(self, iserv):
        self.iserv = iserv
    
    ### ============================= ###
    ### Management einer Projektwoche ###
    ### ============================= ###

    def set_Kurswahl_id(self, Nummer):
        """Lege die Id der Kurswahl fest"""
        self.Kurswahl = str(Nummer)
        
        
    def read_Kurswahl_Angebote(self):
        """Liest Angebote aus der vorgesehenen Excel-Datei in die Datenbank ein."""
        self.message = ""
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        
        c.execute("DROP TABLE IF EXISTS Projekte")
        sql = "CREATE TABLE Projekte (Titel TEXT, Lehrer TEXT, Beschreibung TEXT,"
        sql += "von INTEGER, bis INTEGER, Teilnehmer INTEGER, "
        sql += "Kosten TEXT, Bemerkungen TEXT, eMail TEXT, Id TEXT)"
        c.execute(sql)
        
        
        wb = load_workbook(filename = self.Kurswahl_Angebote)
        ws = wb.active
        max = ws.max_row + 1
        self.message += "Lese " + str(max-2) + " Projekte ein:"
        for i in range (2,max):
            
            Titel = ws.cell(row = i, column = 1).value
            Lehrer = ws.cell(row = i, column = 2).value
            von = int(ws.cell(row = i, column = 3).value)
            bis = int(ws.cell(row = i, column = 4).value)
            Teilnehmer = int(ws.cell(row = i, column = 5).value)
            
            ### Diese Daten werden in dieser Datei nicht erhoben. ###
            Beschreibung = ""
            Kosten = ""
            eMail = ""
            Id = ""
            Bemerkungen = ""
            
            sql = "INSERT INTO Projekte (Titel, Lehrer, Beschreibung, von, bis, Teilnehmer, "
            sql += "Kosten, eMail, Id, Bemerkungen) VALUES (?,?,?,?,?,?,?,?,?,?)"
            c.execute(sql, (Titel, Lehrer, Beschreibung, von, bis, Teilnehmer, Kosten, eMail, Id, Bemerkungen))
            conn.commit()
            
         
        ### Vergabe der IDs in alphabetischer Reihenfolge ###  
        c.execute("SELECT Titel, von, bis FROM Projekte ORDER BY Titel")
        Data = c.fetchall()
        
        i = 0
        for row in Data:
            Titel = row[0]
            von = str(row[1])
            bis = str(row[2])
            i = i + 1
            if (i < 10):
                Id = "P0" + str(i)
            else:
                Id = "P" + str(i)
                
            IdTitel = Id + ": " + row[0] + " (Jg. " + von + "-" + bis +")"
            
            c.execute("UPDATE Projekte SET Id = ?, Titel = ? WHERE Titel = ?", (Id, IdTitel, Titel))
            conn.commit()
            
            self.message += "\n- " + IdTitel
            
        
        
        c.close
        conn.close()

    def export_Kurswahl_Angebote(self):
        """Exportiert die Kurswahl-Angebote als HTML-Liste"""
        self.message =""
        conn = sqlite3.connect(self.database)
        c = conn.cursor() 
        
        c.execute("SELECT Titel, von, bis FROM Projekte ORDER BY Titel")
        Data = c.fetchall()
        
        exportDatei = self.data_dir + "/Projektliste.html"
        file = open(exportDatei, "w")
        file.write("\n<h1>Liste aller Projekte</h1>\n<table>")
        i = 0
        for row in Data:
            i = i + 1
            Titel = row[0]
            von = int(row[1])
            bis = int(row[2])
            TitelJg = Titel + "(Klasse " + str(von) + " bis " + str(bis) + ")"
            if (i % 2 == 1):
                bgColor = "style='background-color:lightyellow'"
            else: 
                bgColor = "style='background-color:white'"
            file.write("\n<tr "+ bgColor + "><td>"+ TitelJg + "</td></tr>")
            
        file.write("\n</table>")
        file.close()
        c.close
        conn.close()
        
    def add_Kurswahl_option(self, Titel, Id, Beschreibung, Teilnehmer):
        """Trägt einen Kurs als Wahloption ein, erfordert Titel, Id, Beschreibung und die maximale Teilnehmerzahl.

           Die Nummer der Kurswahl muss vorher mit set_Kurswahl_Nummer eingestellt werden.
        """
        self.browser.get(self.iserv_url + "courseselection/admin/" + self.Kurswahl_id + "/choice/add")
        max = str(Teilnehmer)
        if (int(Teilnehmer) < 10):
            min = "5"
        else:
            min = "10"

        input = self.browser.find_element_by_name("choice[name]")
        input.send_keys(Titel)

        input = self.browser.find_element_by_name("choice[abbr]")
        input.send_keys(Id)

        input = self.browser.find_element_by_name("choice[min]")
        input.send_keys(min)

        input = self.browser.find_element_by_name("choice[max]")
        input.send_keys(max)

        input = self.browser.find_element_by_name("choice[text]")
        input.send_keys(Beschreibung)

        input = self.browser.find_element_by_name("choice[actions][submit]")
        input.click()
        
    def enter_project_list(self):
        """Trägt die Projekte aus der Datenbank bei iServ in die Umfrage ein."""
        self.message =""
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        
        self.login()
        self.admin_login()
        self.message += "Login erfolgreich."
        
        c.execute("SELECT Titel, Id, Beschreibung, Teilnehmer FROM Projekte ORDER BY Titel")
        Data = c.fetchall()
        
        for row in Data:
            self.add_Kurswahl_option(row[0], row[1], row[2], row[3])
            
        self.logout()
        self.message += "Eintragungen vorgenommen."
        c.close
        conn.close()
        

    def add_Kurswahl_userchoice(self, Vorname, Name, Projekt):
            """Legt Schülerwahlen fest, erfordert den Namen des Schülers und den Namen des Projektes  """

            self.browser.get(self.iserv_url + "courseselection/admin/" + self.Kurswahl + "/userchoice/add")
            time.sleep(4)
            Name = Vorname + " " + Name
            field = self.browser.find_element_by_id("iserv_courseselection_user_choice_crud_user")
            select = Select(field)
            Done = False
            try:
                select.select_by_visible_text(Name)

            except selenium.common.exceptions.NoSuchElementException:
                print("    ... " + Name + " - nicht eingetragen")

            else:
                field = self.browser.find_element_by_id("iserv_courseselection_user_choice_crud_choice_fixed")
                select = Select(field)
                select.select_by_visible_text(Projekt)

                field = self.browser.find_element_by_id("iserv_courseselection_user_choice_crud_choice_one")
                select = Select(field)
                select.select_by_index(1)

                field = self.browser.find_element_by_id("iserv_courseselection_user_choice_crud_choice_two")
                select = Select(field)
                select.select_by_index(2)

                field = self.browser.find_element_by_id("iserv_courseselection_user_choice_crud_choice_three")
                select = Select(field)
                select.select_by_index(3)

                input = self.browser.find_element_by_name("iserv_courseselection_user_choice_crud[actions][submit]")
                input.click()
                Done = True

            return Done
        
    def get_project_title_by_id(self, Id):
        """Ermittelt den Titel des Projekts anhand der Id"""

        conn = sqlite3.connect(self.database)
        c = conn.cursor()

        c.execute("SELECT Titel FROM Projekte WHERE Id = ?", (Id,))
        Data = c.fetchall()
        c.close
        conn.close()
        if Data:
           title = Data[0][0]
        else:
            title = Id
        return title
        

    def read_Kurswahl_choices(self):
        """Liest die Wahlergebnisse in die Datenbank ein."""
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        
        c.execute("DROP TABLE IF EXISTS Wahl")
        c.execute("CREATE TABLE Wahl (Name TEXT, Klasse TEXT, Wahl_1 TEXT, Wahl_2 TEXT, Wahl_3 TEXT, Vorgabe TEXT, Zuweisung TEXT)")
        
        self.login()
        self.admin_login()
        
        ### fordert zum Download der Datei unter Downloads auf
        self.browser.get(self.iserv_url + "courseselection/admin/" + str(self.Kurswahl_id) + "/export")
        time.sleep(4)
        
        self.logout()
        
        ### ermittelt die neueste CSV-Datei im Ordnder Downloads
        last_date = 0
        latest_file =""
        downloads = os.path.expanduser('~') + '/Downloads'
        
        for file in os.listdir(downloads):
           if (".csv" in file):
               date = os.stat(downloads + "/" + file).st_mtime
               if (last_date < date):
                   latest_file = file
                   last_date = date
                
        WahlFile = downloads + "/" + latest_file ## legt fest, dass die neuest CSV-Datei gelesen wird.
        print("reading " + WahlFile)
        with open(WahlFile, 'rt', encoding="iso8859-1") as csvfile:
            i = 0
          
            Data = csv.reader(csvfile, delimiter=";")
            for row in Data:
                if (i > 0):
                    Vorname = row[0].strip()
                    Name = row[1].strip()
                    Klasse = row[3]
                    Jahrgang = row[4]
                    Wahl_1 = row[6]
                    Wahl_2 = row[8]
                    Wahl_3 = row[10]
                    Vorgabe = row[12]
                    Zuweisung = row[14]
                    Wahlen = [Wahl_1, Wahl_2, Wahl_3]
                
                    c.execute("INSERT INTO Wahl (Name, Klasse, Wahl_1, Wahl_2, Wahl_3, Vorgabe, Zuweisung) VALUES (?,?,?,?,?,?,?)",
                            (Vorname + " " + Name, Klasse, Wahl_1, Wahl_2, Wahl_3, Vorgabe, Zuweisung))
                    conn.commit()
                i = i + 1
            
        c.close
        conn.close()
        
    def export_Teilnehmerlisten(self):
        """Exportiert Teilnehmerlisten """
        
        if (not os.path.isdir(self.data_dir + "/export")):
            os.mkdir(self.data_dir + "/export")
        if (not os.path.isdir(self.data_dir + "/export/Klassen")):
            os.mkdir(self.data_dir + "/export/Klassen")
        if (not os.path.isdir(self.data_dir + "/export/Projekte")):
            os.mkdir(self.data_dir + "/export/Projekte")
                         
        self.message = "\nExportiere Teilnehmerlisten nach " + self.data_dir + "/export"
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        
        c.execute("SELECT DISTINCT Zuweisung FROM Wahl ORDER BY Zuweisung ")
        Data = c.fetchall()
        for project in Data:
            Projekt = project[0]
            Titel = Projekt ### self.get_project_title_by_id(Projekt)
            Datei = open(self.data_dir + "/export/Projekte/" +Projekt + ".txt", "w")
            Datei.write("Projekt " + Titel + "\n\n")
            c.execute("SELECT Name, Klasse FROM Wahl WHERE Zuweisung = ? ORDER BY Zuweisung ", (Projekt,))
            Liste = c.fetchall()
            for student in Liste:
                Datei.write("\n" + student[0] + " (" + student[1] + ")")
            
            Datei.close()
        
        c.execute("SELECT DISTINCT Klasse FROM Wahl ORDER BY Klasse ")
        Data = c.fetchall()
        for row in Data:
            Klasse = row[0]
            Datei = open(self.data_dir + "/export/Klassen/" + Klasse + ".txt", "w")
            Datei.write("Klasse " + Klasse + "\n\n")
            c.execute("SELECT Name, Zuweisung FROM Wahl WHERE Klasse = ? ORDER BY Name", (Klasse,))
            Liste = c.fetchall()
            for student in Liste:
                Titel = student[1] ### self.get_project_title_by_id(student[1])
                Datei.write("\n" + student[0] + "\t\t" + Titel)
            
            Datei.close()
            
        c.close
        conn.close()
        
    
    
    def create_project_groups(self):
        """Erstellt iServ-Gruppen und trägt die Teilnehmer ein."""
    
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        
        c.execute("SELECT DISTINCT Zuweisung FROM Wahl ORDER BY Zuweisung ")
        Data = c.fetchall()
        for project in Data:
            Projekt = project[0]
            Titel = Projekt ### self.get_project_title_by_id(Projekt)
            
            c.execute("SELECT Name, Klasse FROM Wahl WHERE Zuweisung = ? ORDER BY Zuweisung ", (Projekt,))
            Liste = c.fetchall()
            for student in Liste:
                Datei.write("\n" + student[0] + " (" + student[1] + ")")
            
            Datei.close()
    
    
    def read_set_students(self):
        """"Liest die festgelegten Schüler aus einer Datei ein und speichert sie in der Tabelle Festlegungen. """
        print("... lese Festlegungen")
        
        conn = sqlite3.connect(self.database)
        c = conn.cursor()
        c.execute('CREATE TABLE IF NOT EXISTS Festlegungen (Vorname TEXT, Name TEXT, Projekt TEXT, erledigt INTEGER)')
        
        if (os.path.isfile(self.data_dir + "/Festlegungen.xlsx")):
            wb=load_workbook(self.data_dir + "/Festlegungen.xlsx")
            sh= wb.active
            max = sh.max_row + 1
            for i in range(2, max):
                Vorname = sh.cell(row=i,column=1)
                Vorname = Vorname.value
                Name = sh.cell(row=i,column=2)
                Name = Name.value
                Projekt = sh.cell(row=i, column=3)
                Projekt = Projekt.value
                
                print(Name+ " " + Vorname + " " + Projekt)
                
                if (i > 0 and Vorname != "" and Name != "" and Projekt != ""):
                    c.execute("SELECT * FROM Festlegungen WHERE Vorname = ? AND Name =?", (Vorname, Name))
                    Data = c.fetchall()
                    if (not Data):
                        c.execute("INSERT INTO Festlegungen (Vorname, Name, Projekt, Erledigt) VALUES (?,?,?,?)",
                              (Vorname, Name, Projekt, 0))
                        conn.commit()
         
        else:
            print("Keine Festlegungen gefunden.\n... Lege eine Excel-Datei an: " + self.data_dir + "/Festlegungen.xlsx")
            wb=Workbook()
            sh=wb.active
            sh['A1'] = "Name" 
            sh['B1'] = "Vorname" 
            sh['C1'] = "Projekt Id" 
            sh['D1'] = "erledigt?"
            sh['E1'] = "Lehrkraft" 
            wb.save(self.data_dir + "/Festlegungen.xlsx")

        c.close
        conn.close()
        
    def add_set_students(self):
        """Fügt die festgelegten Schüler ihren Gruppen zu."""
        self.message = ("Feste Zuordnung von Schülern zu Projekten\n")

        self.read_set_students()

        """        
        conn = sqlite3.connect(self.database)
        c = conn.cursor()

        c.execute("SELECT * FROM Festlegungen ")

        Data = c.fetchall()

        for row in Data:
            Projekt = self.get_project_title_by_id(row[2])
            Vorname = row[0]
            Name = row[1]
            if (int(row[3]) == 0):
                if (iServ.add_Kurswahl_userchoice(Vorname, Name, Projekt)):
                    c.execute("UPDATE Festlegungen SET Erledigt = ? WHERE Name = ? AND Vorname = ?",
                          (1, Name, Vorname))
                    conn.commit()

        iServ.logout()
        """
        
        
    def get_student_id(self, vName):
        """Ermittelt die Schüler-Id aufgrund des Namens"""
        Id = ""
        aName = vName.split(", ")
        vNachname = aName[0]
        vVorname  = aName[1]
        conn = sqlite3.connect(self.database)
        c = conn.cursor()

        c.execute("SELECT Vorname, Id FROM Schueler WHERE Name = ?",
                        (vNachname,))
                
        Data = c.fetchall()
        if Data:
            for row in Data:
                vNameDB = row[0]
                if (re.search(vNameDB, vName)):
                    Id = row[1]
        c.close
        conn.close
    
        return str(Id)        
    
    
        
        
###########################




if __name__ == '__main__':
    print()
    print()

    print("Projektwoche")
    print()

    iServ = iservObject()
    ProWo = ProjektwocheObject(iServ)
    

  
