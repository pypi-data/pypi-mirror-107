import pandas as pd
import openpyxl
from openpyxl.utils.cell import (
    get_column_letter,
    column_index_from_string,
    coordinate_from_string,
)
from openpyxl.utils.dataframe import dataframe_to_rows


# pylint: disable=too-few-public-methods
class PandasExcelWriter:
    def __init__(self, wb):
        assert isinstance(wb, openpyxl.Workbook)
        self.wb = wb

    # pylint: disable=too-many-locals,too-many-arguments
    def write(
        self,
        data: pd.DataFrame,
        ws_name: str,
        node: str,
        header: bool = True,
        index: bool = True,
    ) -> None:
        if not isinstance(data, pd.DataFrame):
            try:
                data = data.to_frame()
            except AttributeError as e:
                raise ValueError(
                    '`data` argument must be a pandas DataFrame or Series, '
                    + 'instead got: %s' % type(data)
                ) from e

        wb = self.wb
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
        else:
            ws = wb.create_sheet(ws_name)
        height, width = data.shape
        if index:
            if isinstance(data.index, pd.MultiIndex):
                width += len(data.index.levels)
            else:
                width += 1
            height += 1  # dataframe_to_rows adds a FrozenList for the index names, even if unnamed
        if header:
            if isinstance(data.columns, pd.MultiIndex):
                height += len(data.columns.levels)
        node_col, node_row = coordinate_from_string(node)
        end_col = get_column_letter(column_index_from_string(node_col) + width - 1)
        end_row = node_row + height
        xl_range = ws[f'{node}:{end_col}{end_row}']
        df_rows = dataframe_to_rows(data, index=index, header=header)
        for df_row, xl_row in zip(df_rows, xl_range):
            for df_data, xl_cell in zip(df_row, xl_row):
                xl_cell.value = df_data
