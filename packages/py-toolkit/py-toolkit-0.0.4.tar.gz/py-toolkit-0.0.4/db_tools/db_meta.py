from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from peewee import Proxy
from playhouse.db_url import connect


class MetaExporter:
    headers: list = ["db_name", "table_name", "column_name", "column_comment", "column_type", "char_length",
                     "is_primary", "is_nullable", "comment"]
    row_mapper: dict = {}
    db = None
    _database = None

    def __init__(self, db_url: str, headers=None, row_mapper=None):
        if not db_url:
            raise ValueError("Database connection config should not be None")

        cls = self.__class__
        cls.init_db(db_url)
        if headers:
            cls.headers = headers

        if row_mapper:
            cls.row_mapper = row_mapper

    @classmethod
    def init_db(cls, db_url):
        cls.db = Proxy()
        cls.db.initialize(connect(db_url))
        cls._database = cls._parse_database(db_url)

    @classmethod
    def _parse_database(cls, db_url: str):
        start = db_url.rindex("/")
        end = db_url.rfind("?")
        if end < 1:
            end = len(db_url)
        return db_url[start + 1:end]

    def export(self, file_name: str = "tables.xlsx"):
        wb, ws = self._init_xlsx()

        schema_meta = self.fetch_schema_meta()
        self._append_schema_meta_to_file(ws, schema_meta)

        wb.save(file_name)

    def _apply_mappers(self, row):
        if self.row_mapper:
            func = self.row_mapper.get(row[2])

            r = list(row)
            if func:
                r[3] = func(row)
            return r
        return row

    def _append_schema_meta_to_file(self, ws: Worksheet, table_info: list):
        start_table_row_idx = 2
        end_table_row_idx = 1
        pre_table = table_info[0][1] if table_info else None
        to_merge = []
        ft = Font(name='Arial', size=16)
        alignment = Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=2,
        )
        for r in table_info:

            ws.append(self._apply_mappers(r))
            cur_table = r[1]
            end_table_row_idx += 1
            col1 = ws[f"A{end_table_row_idx}"]
            col2 = ws[f"B{end_table_row_idx}"]
            col1.font = ft
            col1.alignment = alignment
            col2.font = ft
            col2.alignment = alignment
            if cur_table != pre_table:
                pre_table = cur_table
                to_merge.append((start_table_row_idx, end_table_row_idx - 1))
                start_table_row_idx = end_table_row_idx

        to_merge.append((start_table_row_idx, end_table_row_idx))
        for start, end in to_merge:
            ws.merge_cells(start_row=start, end_row=end, start_column=2, end_column=2)
        ws.merge_cells(start_row=2, start_column=1, end_row=len(table_info) + 1, end_column=1)

    @classmethod
    def _init_xlsx(cls):
        wb = Workbook()
        wb._fonts[0] = Font(name="Arial", sz=14, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
        ws = wb.active
        ws.title = "Tables"
        ws.append(cls.headers)
        return wb, ws

    @classmethod
    def fetch_schema_meta(cls):
        sql = (
            f"SELECT c.TABLE_SCHEMA, c.TABLE_NAME, c.COLUMN_NAME, c.COLUMN_COMMENT, c.COLUMN_TYPE, "
            f"   c.CHARACTER_MAXIMUM_LENGTH, "
            f"   IF(c.COLUMN_KEY='PRI','Yes',' ') AS 'is_primary', "
            f"   IF(c.IS_NULLABLE='YES','Yes',' ') AS 'is_nullable', "
            f"   COLUMN_COMMENT "
            f"FROM information_schema.columns c "
            f"LEFT JOIN information_schema.key_column_usage kcu ON kcu.TABLE_SCHEMA='{cls._database}' "
            f"         AND c.TABLE_NAME = kcu.TABLE_NAME AND c.COLUMN_NAME = kcu.COLUMN_NAME "
            f"WHERE c.TABLE_SCHEMA='{cls._database}'"
        )
        res = cls.db.execute_sql(sql)
        return res.fetchall()
