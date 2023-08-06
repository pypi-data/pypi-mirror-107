
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, pyqtSlot
from PyQt5.QtGui import *
import sys
import os.path
import numpy as np
from openpyxl import Workbook, load_workbook
from SocketFactory.utilities import utility, file_handler
from SocketFactory.outcomes import slices

class SlicerGui(QWidget):

    def __init__(self):                                                        
        super().__init__()
        self.setWindowTitle('Slicer')
        self.initGUI()
        self.apex = "a"
        self.apex_offset = 40
        self.insert_first_page()

    def initGUI(self):

        self.width = 700
        self.height = 300
        self.resize(self.width, self.height)
        self.stacked_widget = QStackedWidget()
        vbox = QVBoxLayout()
        vbox.addWidget(self.stacked_widget)
        self.setLayout(vbox)

    def insert_first_page(self, index=-1):

        self.main_widget = QWidget()
        grid = QGridLayout()
        grid.setSpacing(10)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 2)
        self.main_widget.setLayout(grid)
        grid.addWidget(self.mesh_selection(), 0, 0)
        grid.addWidget(self.set_parameters(), 0, 1)
        self.stacked_widget.insertWidget(index, self.main_widget)
        self.show()

    def mesh_selection(self):

        groupBox = QGroupBox("Files selection")
        self.mesh_paths = []
        self.mesh_label = QLabel("")
        mesh_btn = QPushButton("Load mesh files")
        mesh_btn.clicked.connect(self.choose_meshes)

        self.landmark_paths = []
        self.landmark_label = QLabel("")
        landmark_btn = QPushButton("Load landmark files")
        landmark_btn.clicked.connect(self.choose_landmarks)

        vbox = QVBoxLayout()
        vbox.addWidget(self.mesh_label)
        vbox.addWidget(mesh_btn)
        vbox.addWidget(self.landmark_label)        
        vbox.addWidget(landmark_btn)
        groupBox.setLayout(vbox)
        return groupBox

    @pyqtSlot()
    def choose_meshes(self):
        self.mesh_paths, _ = QFileDialog.getOpenFileNames(self, 'Open file', filter = "Meshes (*.stl *.ply)")
        self.mesh_paths.sort()
        final_label = ''
        for n in self.mesh_paths:
             final_label += n.split("/")[-1] + '\n'
        self.mesh_label.setText(final_label)
        return

    @pyqtSlot()
    def choose_landmarks(self):
        self.landmark_paths, _ = QFileDialog.getOpenFileNames(self, 'Open file', filter = "Text (*.txt)")
        self.landmark_paths.sort()
        final_label = ''
        for n in self.landmark_paths:
             final_label += n.split("/")[-1] + '\n'
        self.landmark_label.setText(final_label)
        return
 
    def set_parameters(self):
        groupBox = QGroupBox("Mode and parameters selection")
        vbox = QVBoxLayout()

        self.radiobutton_lmk = QRadioButton("Landmarks mode")
        self.radiobutton_lmk.setChecked(True)
        self.radiobutton_um = QRadioButton("Specify number slices mode")
        self.radiobutton_um.setChecked(False)
        self.radiobutton_im = QRadioButton("Specify interval mode")
        self.radiobutton_im.setChecked(False)

        grid_layout1 = QGridLayout()
        slices_number_label = QLabel("Slices number: ")
        self.slices_number = QSpinBox()
        grid_layout1.addWidget(slices_number_label, 0, 0)
        grid_layout1.addWidget(self.slices_number, 0, 1)
        start_point_label1 = QLabel("Starting point: ")
        self.start_point1 = QLineEdit()
        grid_layout1.addWidget(start_point_label1, 1, 0)
        grid_layout1.addWidget(self.start_point1, 1, 1)
        end_point_label1 = QLabel("Ending point: ")
        self.end_point1 = QLineEdit()
        grid_layout1.addWidget(end_point_label1, 2, 0)
        grid_layout1.addWidget(self.end_point1, 2, 1)

        grid_layout2 = QGridLayout()
        interval_width_label = QLabel("Interval width (mm): ")
        self.interval_width = QSpinBox()
        grid_layout2.addWidget(interval_width_label, 0, 0)
        grid_layout2.addWidget(self.interval_width, 0, 1)
        start_point_label2 = QLabel("Starting point (distal): ")
        self.start_point2 = QLineEdit()
        grid_layout2.addWidget(start_point_label2, 1, 0)
        grid_layout2.addWidget(self.start_point2, 1, 1)
        end_point_label2 = QLabel("Ending point (proximal): ")
        self.end_point2 = QLineEdit()
        grid_layout2.addWidget(end_point_label2, 2, 0)
        grid_layout2.addWidget(self.end_point2, 2, 1)

        run_btn = QPushButton("Create slices")
        run_btn.setStyleSheet("background-color : lightblue")
        run_btn.clicked.connect(self.run)

        vbox.addWidget(self.radiobutton_lmk)
        vbox.addWidget(self.radiobutton_um)
        vbox.addLayout(grid_layout1)
        vbox.addWidget(self.radiobutton_im)
        vbox.addLayout(grid_layout2)
        vbox.addWidget(run_btn)
        groupBox.setLayout(vbox)
        return groupBox

    def handle_landmarks(self, m_path, l_path):
        
        mesh = utility.read(m_path)
        lmk_dict = file_handler.read_points(l_path)
        return mesh, lmk_dict

    def handle_starting_ending_point(self, m_path, l_path, starting_point, ending_point):
        mesh, lmk_dict = self.handle_landmarks(m_path, l_path)
        lmk_z_heights = utility.query_dict(lmk_dict, [starting_point, ending_point, self.apex], mesh.GetPoints())[:, 2]
        starting_point_height = lmk_z_heights[0]
        ending_point_heigth = lmk_z_heights[1]
        apex40_height = lmk_z_heights[2] + self.apex_offset
        return mesh, starting_point_height, ending_point_heigth, apex40_height

    @pyqtSlot()
    def run(self):

        # Init output measures file
        wb = Workbook()
        ws = wb.active

        if (self.radiobutton_lmk.isChecked()) :
            for m_path, l_path in zip(self.mesh_paths, self.landmark_paths) :
                mesh, lmk_dict = self.handle_landmarks(m_path, l_path)
                lmk_names = lmk_dict.keys()
                lmk_z_heights = utility.query_dict(lmk_dict, lmk_names, mesh.GetPoints())[:, 2]
                self.slices_heights = dict(zip(lmk_z_heights, lmk_names))
                min_index = np.where(lmk_z_heights == np.min(lmk_z_heights))
                # Add apex+offset
                min_height = lmk_z_heights[min_index][0] + self.apex_offset
                self.slices_heights[min_height] = ("a+" + str(self.apex_offset))
                # Order lankmarks by name
                self.slices_heights = {k: v for k, v in sorted(self.slices_heights.items(), key=lambda item: item[1])}
                ws, dir_path = slices.create_slices(mesh, self.slices_heights, m_path, ws)

        elif (self.radiobutton_um.isChecked()) :
            starting_point = self.start_point1.text().lower()
            ending_point =  self.end_point1.text().lower()
            slices_number = self.slices_number.value()

            for m_path, l_path in zip(self.mesh_paths, self.landmark_paths) :
                mesh, starting_point_height, ending_point_heigth, apex40_height = self.handle_starting_ending_point(m_path, l_path, starting_point, ending_point)
                heights = np.linspace(starting_point_height, ending_point_heigth, slices_number)
                heights_percentage = np.linspace(0, 100, len(heights))
                self.slices_heights = dict(zip(heights, map(lambda x : str(round(x, 1)) + "%", heights_percentage)))
                self.slices_heights[apex40_height] = ("a+" + str(self.apex_offset))
                ws, dir_path = slices.create_slices(mesh, self.slices_heights, m_path, ws)

        else :
            starting_point = self.start_point2.text().lower()
            ending_point = self.end_point2.text().lower()
            slices_interval = self.interval_width.value()

            for m_path, l_path in zip(self.mesh_paths, self.landmark_paths) :
                 mesh, starting_point_height, ending_point_heigth, apex40_height = self.handle_starting_ending_point(m_path, l_path, starting_point, ending_point)
                 self.slices_heights = np.arange(starting_point_height, ending_point_heigth, slices_interval)

                 # Add apex+offset and specified end point
                 self.slices_heights = np.append(self.slices_heights, [apex40_height, ending_point_heigth])
                 ws, dir_path = slices.create_slices(mesh, self.slices_heights, m_path, ws)

        self.wb_path = os.path.join(dir_path, "Measures_slices.xlsx")
        wb.save(self.wb_path)

        new_index = self.stacked_widget.currentIndex() + 1 
        self.stacked_widget.insertWidget(new_index, self.insert_second_page())
        self.stacked_widget.setCurrentIndex(new_index)

    def insert_second_page(self):
        second_page = QWidget()
        layout = QVBoxLayout()
        self.mesh_paths = []
        self.mesh_label = QLabel("")
        select_other_meshes_btn = QPushButton("Select meshes")
        select_other_meshes_btn.clicked.connect(self.choose_meshes)
        apply_slicing_btn = QPushButton("Apply same slicing")
        apply_slicing_btn.clicked.connect(self.apply_same_slicing)
        layout.addWidget(self.mesh_label)
        layout.addWidget(select_other_meshes_btn)
        layout.addWidget(apply_slicing_btn)
        second_page.setLayout(layout)
        return second_page

    def apply_same_slicing(self):
        """
        Apply same slicing used for last mesh to all selected.
        """

        wb = load_workbook(filename = self.wb_path)
        ws = wb.active
        for mesh_path in self.mesh_paths:
            mesh = utility.read(mesh_path)
            ws, _ = slices.create_slices(mesh, self.slices_heights, mesh_path, ws)
        wb.save(self.wb_path)